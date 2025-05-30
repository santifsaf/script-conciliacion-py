import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import argparse

def cargar_datos(path_sistema, path_afip):
    """Carga los archivos de sistema y AFIP como DataFrames."""
    df_sis = pd.read_excel(path_sistema, header=0)
    df_afip = pd.read_excel(path_afip, header=1)
    return df_sis, df_afip

def limpiar_datos(df_sis, df_afip):
    """
    Limpia los datos para la conciliación:
    - Separa punto de venta y número de factura.
    - Convierte importes y datos clave a numérico.
    - Renombra columnas.
    """
    df_sis[['punto_venta_sistema', 'nro_factura_sistema']] = df_sis['Numero'].str.split('-', expand=True)
    df_sis['punto_venta_sistema'] = pd.to_numeric(df_sis['punto_venta_sistema'], errors='coerce')
    df_sis['nro_factura_sistema'] = pd.to_numeric(df_sis['nro_factura_sistema'], errors='coerce')


    df_afip['punto_venta_afip'] = pd.to_numeric(df_afip['Punto de Venta'], errors='coerce')
    df_afip['nro_factura_afip'] = pd.to_numeric(df_afip['Numero Desde'], errors='coerce')


    df_sis['importe_sistema'] = pd.to_numeric(
        df_sis['Total'].astype(str).str.replace(r'[^0-9.,-]', '', regex=True).str.replace(',', ''), errors='coerce')
    df_afip['importe_afip'] = pd.to_numeric(
        df_afip['Imp. Total'].astype(str).str.replace(r'[^0-9.,-]', '', regex=True).str.replace(',', ''), errors='coerce')


    df_sis.rename(columns={'Fecha': 'fecha_sistema', 'Proveedor': 'proveedor'}, inplace=True)
    df_afip.rename(columns={'Fecha': 'fecha_afip', 'Denominación Emisor': 'emisor'}, inplace=True)
    df_afip.columns = df_afip.columns.str.strip()

    return df_sis, df_afip

def clasificar(row):
    """
    Clasifica cada comprobante según la comparación de datos clave e importes.
    """
    if pd.notna(row['punto_venta_afip']) and pd.notna(row['nro_factura_afip']) and pd.notna(row['punto_venta_sistema']) and pd.notna(row['nro_factura_sistema']):
        if pd.notna(row['importe_afip']) and pd.notna(row['importe_sistema']):
            if abs(row['importe_afip'] - row['importe_sistema']) < 0.01:
                return 'Coincidencia'
            else:
                return 'Importe no coincide'
        return 'Importe faltante'
        # Solo en sistema
    elif pd.isna(row['punto_venta_afip']) or pd.isna(row['nro_factura_afip']):
        return 'Solo en sistema'
        # Solo en AFIP
    elif pd.isna(row['punto_venta_sistema']) or pd.isna(row['nro_factura_sistema']):
        return 'Solo en AFIP'
    return 'Otro'

def comparar(df_sis, df_afip):
    """
    Realiza el merge entre sistema y AFIP y clasifica cada comprobante.
    Devuelve un DataFrame con todas las filas y su condición.
    """
    cols_afip = ['fecha_afip', 'emisor', 'punto_venta_afip', 'nro_factura_afip', 'importe_afip']
    cols_sis = ['fecha_sistema', 'proveedor', 'punto_venta_sistema', 'nro_factura_sistema', 'importe_sistema']

    # merge de ambos DataFrames por punto de venta y número de factura
    df_merged = pd.merge(
        df_afip[cols_afip], df_sis[cols_sis],
        left_on=['punto_venta_afip', 'nro_factura_afip'],
        right_on=['punto_venta_sistema', 'nro_factura_sistema'],
        how='outer'
    )

    # Elimina filas totalmente vacías 
    df_merged = df_merged[
        df_merged[['punto_venta_afip', 'nro_factura_afip', 'punto_venta_sistema', 'nro_factura_sistema']].notna().any(axis=1)
    ]

    # Clasifica cada comprobante
    df_merged['condicion'] = df_merged.apply(clasificar, axis=1)
    return df_merged

def exportar(df, df_afip_original, filename='resultados_facturas.xlsx'):
    """
    Exporta los resultados a un archivo Excel con hojas separadas por tipo de coincidencia.
    También colorea la columna de importe según la condición.
    """
    colores = {
        "Solo en AFIP": "FFCCCB",            
        "Importe no coincide": "FFD700",     
        "Coincidencia": "98FB98",            
    }

    with pd.ExcelWriter(filename) as writer:
        # Exporta comprobantes en solapas diferentes
        solo_en_afip = df[df['condicion'] == 'Solo en AFIP']
        solo_en_afip.to_excel(writer, sheet_name='Solo en AFIP', index=False)
        df[df['condicion'] == 'Solo en sistema'].to_excel(writer, sheet_name='Solo en el SISTEMA', index=False)
        df[df['condicion'] == 'Importe no coincide'].to_excel(writer, sheet_name='Coincidencia parcial', index=False)
        df[df['condicion'] == 'Coincidencia'].to_excel(writer, sheet_name='Coinciden', index=False)


        df_afip_con_condicion = pd.merge(
            df_afip_original,
            df[['punto_venta_afip', 'nro_factura_afip', 'condicion']],
            on=['punto_venta_afip', 'nro_factura_afip'],
            how='left'
        )

        columnas_finales = [
            'fecha_afip', 'emisor', 'punto_venta_afip', 'nro_factura_afip', 'importe_afip', 'condicion'
        ]
        df_afip_con_condicion = df_afip_con_condicion[columnas_finales]
        df_afip_con_condicion.rename(columns={
            'fecha_afip': 'Fecha',
            'emisor': 'Emisor',
            'punto_venta_afip': 'Punto de venta',
            'nro_factura_afip': 'Nro de factura',
            'importe_afip': 'Imp total',
            'condicion': 'Condición'
        }, inplace=True)

        df_afip_con_condicion.to_excel(writer, sheet_name='Mis Comprobantes Recibidos', index=False)

    # Colorear la columna de importe según la condición
    wb = load_workbook(filename)
    ws = wb["Mis Comprobantes Recibidos"]
    for idx, cell in enumerate(ws[1], 1):
        if str(cell.value).strip().lower() == 'imp. total':
            col = idx
            for fila in range(2, ws.max_row + 1):
                condicion = ws.cell(row=fila, column=ws.max_column).value
                color = colores.get(condicion)
                if color:
                    ws.cell(row=fila, column=col).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            break
    wb.save(filename)

def main():
    parser = argparse.ArgumentParser(description="Conciliación de facturas entre sistema y AFIP.")
    parser.add_argument('--sistema', default='Facturas en sistema.xlsx', help='Archivo de facturas del sistema')
    parser.add_argument('--afip', default='Facturas en afip.xlsx', help='Archivo de facturas de AFIP')
    args = parser.parse_args()

    df_sis, df_afip = cargar_datos(args.sistema, args.afip)
    df_sis, df_afip = limpiar_datos(df_sis, df_afip)
    df_resultado = comparar(df_sis, df_afip)
    exportar(df_resultado, df_afip)
    print("Comparación finalizada. Archivo generado: resultados_facturas.xlsx")

if __name__ == '__main__':
    main()